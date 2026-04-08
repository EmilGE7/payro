import os
import io
import uuid
import time
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, redirect, url_for, request, flash, session, send_file, Response
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from sqlalchemy import text
from sqlalchemy.dialects.postgresql import UUID
from werkzeug.security import generate_password_hash, check_password_hash
from dotenv import load_dotenv
from fpdf import FPDF
from openai import OpenAI
from flask_caching import Cache
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Load environment variables
load_dotenv()

# --- Shared OpenAI Helper ---
_openai_client = None

def get_openai_client():
    global _openai_client
    if _openai_client is None:
        key = os.environ.get("AI_API_KEY") # Shared OpenAI Key
        if key:
            _openai_client = OpenAI(api_key=key)
    return _openai_client

def ask_openai(system_prompt: str, user_content: str) -> str:
    """Shared helper for all OpenAI features. Returns empty string on error."""
    client = get_openai_client()
    if not client:
        return ""
    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_content}
            ],
            max_tokens=1024
        )
        return response.choices[0].message.content
    except Exception:
        return ""

# --- Database & Models ---
db = SQLAlchemy()

class User(db.Model, UserMixin):
    __tablename__ = 'users'
    id = db.Column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    name = db.Column(db.String(80), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False) # admin, hr, accounting, employee
    profile = db.relationship('EmployeeProfile', backref='user', uselist=False)

class Department(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    employees = db.relationship('EmployeeProfile', backref='dept')

class EmployeeProfile(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(UUID(as_uuid=True), db.ForeignKey('users.id'), nullable=False)
    dept_id = db.Column(db.Integer, db.ForeignKey('department.id'))
    job_title = db.Column(db.String(100))
    joining_date = db.Column(db.DateTime, default=datetime.utcnow)
    contact = db.Column(db.String(20))
    address = db.Column(db.Text)
    salary_structure = db.relationship('SalaryStructure', backref='profile', uselist=False)

class SalaryStructure(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    profile_id = db.Column(db.Integer, db.ForeignKey('employee_profile.id'), nullable=False)
    base_salary = db.Column(db.Float, default=0.0)
    allowances = db.Column(db.Float, default=0.0)
    deductions = db.Column(db.Float, default=0.0)

class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(UUID(as_uuid=True), db.ForeignKey('users.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    status = db.Column(db.String(20), nullable=False)

class LeaveRequest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(UUID(as_uuid=True), db.ForeignKey('users.id'), nullable=False)
    start_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date, nullable=False)
    reason = db.Column(db.Text)
    status = db.Column(db.String(20), default='Pending')
    user = db.relationship('User', backref='leave_requests')

class PayrollRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(UUID(as_uuid=True), db.ForeignKey('users.id'), nullable=False)
    month = db.Column(db.Integer, nullable=False)
    year = db.Column(db.Integer, nullable=False)
    net_amount = db.Column(db.Float, nullable=False)
    paid_date = db.Column(db.DateTime, default=datetime.utcnow)
    status = db.Column(db.String(20), default='Paid')
    user = db.relationship('User', backref='payroll_records', lazy=True)

class SalaryChangeRequest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(UUID(as_uuid=True), db.ForeignKey('users.id'), nullable=False)
    requested_by = db.Column(UUID(as_uuid=True), db.ForeignKey('users.id'), nullable=False)
    change_type = db.Column(db.String(20), nullable=False) # 'increment', 'decrement'
    amount = db.Column(db.Float, nullable=False)
    reason = db.Column(db.Text)
    status = db.Column(db.String(20), default='Pending') # Pending, Approved, Rejected
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    user = db.relationship('User', foreign_keys=[user_id], backref='salary_requests')
    requester = db.relationship('User', foreign_keys=[requested_by])

class Notification(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(UUID(as_uuid=True), db.ForeignKey('users.id'), nullable=False)
    message = db.Column(db.Text, nullable=False)
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class SalaryHistory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(UUID(as_uuid=True), db.ForeignKey('users.id'), nullable=False)
    old_salary = db.Column(db.Float)
    new_salary = db.Column(db.Float)
    changed_at = db.Column(db.DateTime, default=datetime.utcnow)
    changed_by = db.Column(db.String(80))

# --- Business Logic: AI Engine ---
def analyze_payroll_data(db_session, user_prompt=None):
    api_key = os.environ.get("AI_API_KEY")
    total_payroll_count = db_session.query(PayrollRecord).count()
    total_users = User.query.count()
    pending_leaves = LeaveRequest.query.filter_by(status='Pending').count()
    context = f"System State: {total_users} employees, {total_payroll_count} payroll records, {pending_leaves} pending leaves."
    if not api_key:
        return f"Offline: Query '{user_prompt}' received. Workforce: {total_users}." if user_prompt else "Analysis: Payroll within normal parameters."
    try:
        client = get_openai_client()
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": "You are an AI Payroll Consultant. Provide a single-sentence executive insight."},
                {"role": "user", "content": f"{context}\n\nQuestion/Prompt: {user_prompt or 'General status insight'}"}
            ],
            max_tokens=80
        )
        return response.choices[0].message.content
    except Exception as e: return f"AI Error: {str(e)}"

# --- Business Logic: PDF Helpers ---
def setup_pdf_fonts(pdf):
    """Safely adds DejaVu fonts if available, otherwise returns helvetica."""
    # Common paths for Render (Ubuntu/Debian)
    font_regular = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
    font_bold = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
    
    if os.path.exists(font_regular) and os.path.exists(font_bold):
        try:
            pdf.add_font('DejaVu', '', font_regular, uni=True)
            pdf.add_font('DejaVu', 'B', font_bold, uni=True)
            return 'DejaVu'
        except Exception:
            pass
    return 'helvetica'

# --- Business Logic: Payslip Generation ---
class PayslipGenerator(FPDF):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.font_family_name = setup_pdf_fonts(self)

    def header(self):
        self.set_font(self.font_family_name, 'B', 20)
        self.set_text_color(99, 102, 241)
        self.cell(0, 10, 'Payro DBMS', ln=True, align='L')
        self.set_font(self.font_family_name, '', 10)
        self.set_text_color(100)
        self.cell(0, 5, 'Executive Financial Document', ln=True, align='L')
        self.ln(10)
    def footer(self):
        self.set_y(-15)
        self.set_font(self.font_family_name, '', 8)
        self.set_text_color(150)
        self.cell(0, 10, f'Page {self.page_no()} | Generated by Payro System', align='C')

def generate_payslip_pdf(payroll_record):
    if not payroll_record or not payroll_record.user:
        pdf = FPDF(); pdf.add_page(); pdf.set_font('helvetica', 'B', 16); pdf.cell(0, 10, "Error: User data missing", ln=True); return pdf.output()
    pdf = PayslipGenerator()
    family = pdf.font_family_name
    pdf.add_page()
    pdf.set_fill_color(248, 250, 252)
    pdf.set_font(family, 'B', 12)
    pdf.cell(0, 10, 'Employee Details', ln=True, fill=True)
    pdf.set_font(family, '', 10)
    pdf.cell(95, 7, f'Name: {payroll_record.user.name}', ln=False)
    pdf.cell(95, 7, f'Email: {payroll_record.user.email}', ln=True)
    pdf.cell(95, 7, f'Month/Year: {payroll_record.month}/{payroll_record.year}', ln=False)
    job_title = payroll_record.user.profile.job_title if payroll_record.user.profile else "N/A"
    pdf.cell(95, 7, f'Designation: {job_title}', ln=True)
    pdf.ln(10)
    pdf.set_font(family, 'B', 12); pdf.cell(0, 10, 'Salary Breakdown', ln=True, fill=True)
    pdf.set_font(family, 'B', 10); pdf.cell(100, 10, 'Description', border=1); pdf.cell(90, 10, 'Amount (INR)', border=1, ln=True, align='R')
    pdf.set_font(family, '', 10)
    ss = payroll_record.user.profile.salary_structure if payroll_record.user.profile else None
    base, allowances, deductions = (ss.base_salary, ss.allowances, ss.deductions) if ss else (0, 0, 0)
    pdf.cell(100, 10, 'Base Salary', border=1); pdf.cell(90, 10, f'₹{base:,.2f}', border=1, ln=True, align='R')
    pdf.cell(100, 10, 'Allowances', border=1); pdf.cell(90, 10, f'₹{allowances:,.2f}', border=1, ln=True, align='R')
    pdf.set_text_color(244, 63, 94); pdf.cell(100, 10, 'Deductions', border=1); pdf.cell(90, 10, f'-₹{deductions:,.2f}', border=1, ln=True, align='R')
    pdf.set_text_color(0); pdf.set_font(family, 'B', 12); pdf.cell(100, 12, 'NET PAYABLE', border=1); pdf.cell(90, 12, f'₹{payroll_record.net_amount:,.2f}', border=1, ln=True, align='R')
    pdf.ln(20); pdf.set_font(family, '', 10); pdf.multi_cell(0, 5, 'Computer-generated document. No signature required.')
    return pdf.output()

# --- App Initialization ---
app = Flask(__name__, static_folder='static', template_folder='templates')
DATABASE_URL = os.environ.get("DATABASE_URL")

if DATABASE_URL:
    if DATABASE_URL.startswith("postgres://"):
        DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql+psycopg2://", 1)
    elif DATABASE_URL.startswith("postgresql://"):
        DATABASE_URL = DATABASE_URL.replace("postgresql://", "postgresql+psycopg2://", 1)
    if "sslmode" not in DATABASE_URL:
        DATABASE_URL += ("&" if "?" in DATABASE_URL else "?") + "sslmode=require"
    DATABASE_URL = DATABASE_URL + "&keepalives=1&keepalives_idle=30"

app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Security: Fallback SECRET_KEY for initial deployment safety
app.config['SECRET_KEY'] = os.environ.get("SECRET_KEY")
if app.config['SECRET_KEY'] is None:
    print("WARNING: SECRET_KEY not found. Using temporary fallback. Set SECRET_KEY in Render Environment!")
    app.config['SECRET_KEY'] = "dev-fallback-key-change-me"

# DB Connection Check
if not DATABASE_URL:
    print("ERROR: DATABASE_URL not found. App will crash on first DB access. Set DATABASE_URL in Render Environment!")

app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    "pool_size": 2, # Reduced for free-tier compatibility
    "max_overflow": 0,
    "pool_timeout": 30,
    "pool_recycle": 300,
    "pool_pre_ping": True,
}

db.init_app(app)
cache = Cache(app, config={'CACHE_TYPE': 'simple'})
login_manager = LoginManager(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    try: return db.session.get(User, user_id)
    except Exception: return None

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        @login_required
        def decorated_function(*args, **kwargs):
            if current_user.role not in roles:
                flash("Access Denied", "danger")
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# --- Routes ---
@app.errorhandler(404)
def not_found(e):
    return redirect(url_for('login'))

@app.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.role == 'employee':
            return redirect(url_for('view_employee_profile', id=current_user.id))
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        user = User.query.filter_by(email=email).first()
        if user and check_password_hash(user.password, password):
            login_user(user)
            if user.role == 'employee':
                return redirect(url_for('view_employee_profile', id=user.id))
            return redirect(url_for('dashboard'))
        return render_template('login.html', error="Invalid credentials")
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    if current_user.role == 'employee':
        return redirect(url_for('view_employee_profile', id=current_user.id))
    now = datetime.now()
    total_employees = User.query.filter_by(role='employee').count()
    latest_pr = PayrollRecord.query.order_by(PayrollRecord.year.desc(), PayrollRecord.month.desc()).first()
    if latest_pr:
        current_payroll = db.session.query(db.func.sum(PayrollRecord.net_amount)).filter(PayrollRecord.month == latest_pr.month, PayrollRecord.year == latest_pr.year).scalar() or 0.0
    else:
        current_payroll = 0.0
    pending_leaves = LeaveRequest.query.filter_by(status='Pending').count()
    # Fetch latest 5 payroll logs, regardless of paid_date existence
    recent_payroll = PayrollRecord.query.order_by(PayrollRecord.year.desc(), PayrollRecord.month.desc()).limit(5).all()
    return render_template('dashboard.html', user=current_user, now=now, total_employees=total_employees, total_payroll=current_payroll, pending_leaves=pending_leaves, recent_activities=recent_payroll)

@app.route('/employees', methods=['GET', 'POST'])
@role_required('admin', 'hr')
def view_employees():
    if request.method == 'POST' and current_user.role in ['admin', 'hr']:
        name, email, password, role, dept_id_raw = request.form.get('name'), request.form.get('email'), request.form.get('password'), request.form.get('role'), request.form.get('dept_id')
        dept_id = int(dept_id_raw) if dept_id_raw and dept_id_raw.isdigit() else None
        try:
            hashed_password = generate_password_hash(password)
            user = User(name=name, email=email, password=hashed_password, role=role)
            db.session.add(user)
            db.session.flush()
            profile = EmployeeProfile(user_id=user.id, dept_id=dept_id, job_title=f"{role.capitalize()} Staff")
            db.session.add(profile)
            db.session.flush()
            db.session.add(SalaryStructure(profile_id=profile.id, base_salary=50000))
            db.session.add(Notification(user_id=user.id, message="Welcome to Payro!"))
            db.session.commit()
            flash(f"Added {name}", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Error: {str(e)}", "danger")
    return render_template('employees.html', employees=User.query.all(), departments=Department.query.all())

@app.route('/employee/<uuid:id>')
@login_required
def view_employee_profile(id):
    if current_user.role not in ['admin', 'hr'] and current_user.id != id:
        flash("Access Denied", "danger")
        return redirect(url_for('dashboard'))
    
    employee = User.query.get_or_404(id)
    recent_leaves = LeaveRequest.query.filter_by(user_id=id).order_by(LeaveRequest.start_date.desc()).limit(5).all()
    recent_payrolls = PayrollRecord.query.filter_by(user_id=id).order_by(PayrollRecord.year.desc(), PayrollRecord.month.desc()).limit(5).all()
    
    return render_template('employee_profile.html', employee=employee, recent_leaves=recent_leaves, recent_payrolls=recent_payrolls)

@app.route('/attendance', methods=['GET', 'POST'])
@login_required
def view_attendance():
    if request.method == 'POST':
        try:
            if current_user.role == 'employee':
                start = datetime.strptime(request.form.get('start_date'), '%Y-%m-%d').date()
                end = datetime.strptime(request.form.get('end_date'), '%Y-%m-%d').date()
                db.session.add(LeaveRequest(user_id=current_user.id, start_date=start, end_date=end, reason=request.form.get('reason')))
                db.session.commit(); flash("Submitted!", "success")
            elif current_user.role in ['hr', 'admin']:
                leave = LeaveRequest.query.get(request.form.get('leave_id'))
                if leave: leave.status = request.form.get('action'); db.session.commit()
        except Exception: flash("Error!", "danger")
    pending = LeaveRequest.query.filter_by(status='Pending').all() if current_user.role in ['admin', 'hr'] else []
    return render_template('attendance.html', pending_leaves=pending, my_leaves=LeaveRequest.query.filter_by(user_id=current_user.id).all())

@app.route('/payroll', methods=['GET', 'POST'])
@role_required('accounting', 'admin')
def view_payroll():
    if request.method == 'POST':
        m, y = int(request.form.get('month')), int(request.form.get('year'))
        if not PayrollRecord.query.filter_by(month=m, year=y).first():
            for ep in EmployeeProfile.query.all():
                ss = ep.salary_structure
                if ss is None:
                    continue
                net = (ss.base_salary or 0) + (ss.allowances or 0) - (ss.deductions or 0)
                db.session.add(PayrollRecord(user_id=ep.user_id, month=m, year=y, net_amount=net))
            db.session.commit()
    records = PayrollRecord.query.order_by(PayrollRecord.year.desc(), PayrollRecord.month.desc()).all()
    latest_rec = records[0] if records else None
    if latest_rec:
        payout = sum(r.net_amount for r in records if r.month == latest_rec.month and r.year == latest_rec.year)
    else:
        payout = 0
    pending_reqs = SalaryChangeRequest.query.filter_by(status='Pending').all()
    return render_template('payroll.html', records=records, total_payout=payout, tax_est=payout*0.15, bonuses_est=payout*0.05, now=datetime.now(), pending_requests=pending_reqs)

@app.route('/salary/request', methods=['POST'])
@role_required('hr', 'admin')
def request_salary_change():
    try:
        user_id = request.form.get('user_id')
        change_type = request.form.get('change_type')
        amount = float(request.form.get('amount'))
        reason = request.form.get('reason')
        
        db.session.add(SalaryChangeRequest(
            user_id=user_id,
            requested_by=current_user.id,
            change_type=change_type,
            amount=amount,
            reason=reason
        ))
        db.session.commit()
        flash("Salary change application submitted to Accounting.", "success")
    except Exception as e:
        flash(f"Error submitting request: {e}", "danger")
    return redirect(url_for('view_employee_profile', id=user_id))

@app.route('/salary/approve', methods=['POST'])
@role_required('accounting', 'admin')
def approve_salary_change():
    try:
        req_id = request.form.get('request_id')
        action = request.form.get('action') # 'Approved' or 'Rejected'
        req_obj = SalaryChangeRequest.query.get(req_id)
        if req_obj and req_obj.status == 'Pending':
            req_obj.status = action
            if action == 'Approved':
                ss = req_obj.user.profile.salary_structure
                old_salary = ss.base_salary
                if req_obj.change_type == 'increment':
                    ss.base_salary += req_obj.amount
                elif req_obj.change_type == 'decrement':
                    ss.base_salary -= req_obj.amount
                
                db.session.add(SalaryHistory(
                    user_id=req_obj.user_id,
                    old_salary=old_salary,
                    new_salary=ss.base_salary,
                    changed_by=current_user.name
                ))
            db.session.commit()
            flash(f"Salary request {action.lower()}.", "success")
    except Exception as e:
        flash(f"Error processing request: {str(e)}", "danger")
    return redirect(url_for('view_payroll'))

@app.route('/api/ai/analyze', methods=['GET', 'POST'])
@login_required
def ai_analyze():
    try:
        prompt = request.json.get('prompt') if request.method == 'POST' else None
        return {"insight": analyze_payroll_data(db.session, prompt)}
    except Exception as e: return {"insight": f"Error: {str(e)}"}, 500

@app.route('/payroll/download/<id>')
@login_required
def download_payslip(id):
    rec = PayrollRecord.query.get(id)
    if not rec or (current_user.role == 'employee' and rec.user_id != current_user.id): return redirect(url_for('dashboard'))
    return send_file(io.BytesIO(generate_payslip_pdf(rec)), mimetype='application/pdf', as_attachment=True, download_name=f'Payslip_{rec.month}_{rec.year}.pdf')

# --- Antigravity Overhaul (OpenAI Edition) ---

@app.route('/ai/monthly-report')
@login_required
def ai_monthly_report():
    try:
        now = datetime.now(); month_name = now.strftime('%B')
        total_payroll = db.session.query(db.func.sum(PayrollRecord.net_amount)).filter(PayrollRecord.month == now.month, PayrollRecord.year == now.year).scalar() or 0
        employee_count = User.query.filter_by(role='employee').count()
        paid_count = PayrollRecord.query.filter_by(month=now.month, year=now.year, status='Paid').count()
        leaves_approved = LeaveRequest.query.filter_by(status='Approved').count()
        leaves_pending = LeaveRequest.query.filter_by(status='Pending').count()
        leaves_rejected = LeaveRequest.query.filter_by(status='Rejected').count()
        
        from datetime import timedelta
        sixty_days_ago = (now - timedelta(days=60)).date()
        no_leave_names = [emp.name for emp in User.query.filter_by(role='employee').all() if not LeaveRequest.query.filter(LeaveRequest.user_id == emp.id, LeaveRequest.status == 'Approved', LeaveRequest.start_date >= sixty_days_ago).first()]
        
        data = {"month": month_name, "year": now.year, "total_payroll_inr": round(total_payroll, 2), "total_employees": employee_count, "employees_paid_this_month": paid_count, "leave_requests": {"approved": leaves_approved, "pending": leaves_pending, "rejected": leaves_rejected}, "employees_no_leave_60_days": no_leave_names[:5]}
        system = "You are an HR analytics assistant for a company using Payro. Write a concise, professional monthly HR narrative in exactly 3-4 sentences. Focus on payroll cost, leave trends, and any risk signals. Write in plain business English. No bullet points. No markdown. Never invent data — use only the numbers provided. End with one forward-looking sentence."
        narrative = ask_openai(system, f"Generate HR narrative: {json.dumps(data)}")
        if not narrative: narrative = f"In {month_name} {now.year}, {employee_count} employees were on payroll with a total payout of ₹{total_payroll:,.2f}. There are {leaves_pending} pending leave requests."
        session['last_narrative'] = narrative; session['last_narrative_month'] = f"{month_name} {now.year}"
        return {"narrative": narrative, "month": f"{month_name} {now.year}"}
    except Exception: return {"error": "AI service temporarily unavailable"}, 503

@app.route('/ai/monthly-report/pdf')
@login_required
def ai_monthly_report_pdf():
    narrative = session.get('last_narrative'); month = session.get('last_narrative_month', 'Report')
    if not narrative: return redirect(url_for('dashboard'))
    pdf = FPDF()
    family = setup_pdf_fonts(pdf)
    pdf.add_page()
    pdf.set_font(family, 'B', 20)
    pdf.set_text_color(99, 102, 241)
    pdf.cell(0, 12, 'Payro - HR Monthly Report', ln=True)
    pdf.set_font(family, '', 11); pdf.set_text_color(100); pdf.cell(0, 7, month, ln=True); pdf.ln(10)
    pdf.set_font(family, '', 12); pdf.set_text_color(30, 30, 30); pdf.multi_cell(0, 8, narrative); pdf.ln(15)
    pdf.set_font(family, '', 9); pdf.set_text_color(150); pdf.cell(0, 5, f'Generated by Payro AI on {datetime.now().strftime("%Y-%m-%d %H:%M")}', ln=True)
    return send_file(io.BytesIO(pdf.output()), mimetype='application/pdf', as_attachment=True, download_name=f'payro-hr-report-{month.replace(" ", "-")}.pdf')

@app.route('/ai/explain-payslip/<int:payroll_id>')
@login_required
def ai_explain_payslip(payroll_id):
    try:
        rec = PayrollRecord.query.get(payroll_id)
        if not rec or (current_user.role == 'employee' and rec.user_id != current_user.id): return {"error": "Unauthorized"}, 403
        ss = rec.user.profile.salary_structure if rec.user.profile else None
        base = ss.base_salary if ss else rec.net_amount * 0.6; alw = ss.allowances if ss else rec.net_amount * 0.2; ded = ss.deductions if ss else rec.net_amount * 0.1
        pf = round(base * 0.12, 2); tds = round(max(0, (base * 12 - 250000) * 0.05 / 12), 2)
        breakdown = {"employee_name": rec.user.name, "month_year": f"{rec.month}/{rec.year}", "gross_salary": round(base + alw, 2), "basic_salary": round(base, 2), "allowances": round(alw, 2), "pf_deduction": pf, "tds_deduction": tds, "other_deductions": round(max(0, ded - pf - tds), 2), "net_salary": round(rec.net_amount, 2), "currency": "INR"}
        system = "You are a friendly HR assistant explaining an employee's payslip in plain, simple English. Write 2-3 sentences total. Explain what each main component means in everyday language. Be warm, clear, and helpful. Never use financial jargon without explaining it. Do not use bullet points."
        explanation = ask_openai(system, f"Explain this payslip to the employee: {json.dumps(breakdown)}")
        if not explanation: explanation = f"Your gross pay this month was ₹{breakdown['gross_salary']:,.2f}, which includes your base salary of ₹{breakdown['basic_salary']:,.2f} plus allowances. After deductions your net take-home is ₹{breakdown['net_salary']:,.2f}."
        return {"explanation": explanation, "breakdown": breakdown}
    except Exception: return {"error": "AI service temporarily unavailable"}, 503

@app.route('/ai/leave-impact/<int:leave_id>')
@login_required
def ai_leave_impact(leave_id):
    try:
        if current_user.role not in ['admin', 'hr']: return {"error": "Unauthorized"}, 403
        leave = LeaveRequest.query.get(leave_id)
        if not leave: return {"error": "Leave not found"}, 404
        delta = (leave.end_date - leave.start_date).days + 1
        leaves_this_year = LeaveRequest.query.filter(LeaveRequest.user_id == leave.user_id, LeaveRequest.status == 'Approved', db.extract('year', LeaveRequest.start_date) == datetime.now().year).count()
        last_leave = LeaveRequest.query.filter(LeaveRequest.user_id == leave.user_id, LeaveRequest.status == 'Approved').order_by(LeaveRequest.end_date.desc()).first()
        days_since_last = (datetime.now().date() - last_leave.end_date).days if last_leave else None
        overlapping = LeaveRequest.query.filter(LeaveRequest.user_id != leave.user_id, LeaveRequest.status == 'Approved', LeaveRequest.start_date <= leave.end_date, LeaveRequest.end_date >= leave.start_date).count()
        context = {"employee_name": leave.user.name, "leave_start": str(leave.start_date), "leave_end": str(leave.end_date), "leave_duration_days": delta, "reason": leave.reason or "Not specified", "leaves_taken_this_year": leaves_this_year, "remaining_annual_leaves": max(0, 24 - leaves_this_year), "days_since_last_approved_leave": days_since_last, "other_employees_on_leave_same_dates": overlapping}
        system = "You are an HR decision assistant. Given a leave request and employee context, write a 2-3 sentence impact summary for the manager. Be factual, neutral, and helpful. Mention team coverage issues if overlapping leaves exist, attendance impact, and leave balance. End with exactly one sentence starting with 'Recommendation:'"
        impact = ask_openai(system, f"Leave request context: {json.dumps(context)}")
        if not impact: impact = f"{leave.user.name} has taken {leaves_this_year} leaves this year. {overlapping} other employee(s) overlap on these dates. Recommendation: Review team coverage before approving."
        return {"impact": impact}
    except Exception: return {"error": "AI service temporarily unavailable"}, 503

_burnout_cache = {}
@app.route('/api/burnout-scores')
@login_required
def api_burnout_scores():
    try:
        global _burnout_cache; now = datetime.now()
        if _burnout_cache.get('timestamp') and (now - _burnout_cache['timestamp']).seconds / 3600 < 24: return {"scores": _burnout_cache['data']}
        from datetime import timedelta; sixty_days_ago = (now - timedelta(days=60)).date()
        employees = User.query.filter_by(role='employee').all(); at_risk_batch = []; all_scores = {}
        for emp in employees:
            leaves_this_year = LeaveRequest.query.filter(LeaveRequest.user_id == emp.id, LeaveRequest.status == 'Approved', db.extract('year', LeaveRequest.start_date) == now.year).count()
            last_leave = LeaveRequest.query.filter(LeaveRequest.user_id == emp.id, LeaveRequest.status == 'Approved').order_by(LeaveRequest.end_date.desc()).first()
            days_since = (now.date() - last_leave.end_date).days if last_leave else None
            all_scores[str(emp.id)] = {"employee_id": str(emp.id), "risk_level": "healthy", "reason": "Regular leave pattern observed."}
            if (days_since is None or days_since > 60) or leaves_this_year < 3:
                at_risk_batch.append({"employee_id": str(emp.id), "name": emp.name, "leaves_taken_this_year": leaves_this_year, "days_since_last_leave": days_since})
        if at_risk_batch:
            system = "You are an HR wellness analyst. For each employee, return ONLY a valid JSON array. Each item must have exactly: employee_id (string), risk_level (one of: healthy, watch, at_risk), reason (one sentence, max 10 words). Return raw JSON array ONLY. No markdown. No explanation. No preamble."
            raw = ask_openai(system, f"Assess these employees: {json.dumps(at_risk_batch)}")
            if raw:
                try:
                    ai_scores = json.loads(raw.strip().lstrip('```json').lstrip('```').rstrip('```').strip())
                    for score in ai_scores:
                        eid = str(score.get('employee_id', ''))
                        if eid in all_scores: all_scores[eid] = score
                except Exception: pass
        result = list(all_scores.values()); _burnout_cache = {'data': result, 'timestamp': now}
        return {"scores": result}
    except Exception: return {"error": "AI service temporarily unavailable"}, 503

# --- Export Routes ---

@app.route('/export/employees')
@role_required('admin', 'hr')
def export_employees():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Employees"
    header_font = Font(bold=True, color="FFFFFF"); header_fill = PatternFill("solid", fgColor="6366F1")
    headers = ["ID", "Name", "Email", "Role", "Department", "Job Title", "Joining Date", "Contact"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h); cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center')
    for row, emp in enumerate(User.query.all(), 2):
        p = emp.profile
        ws.cell(row=row, column=1, value=str(emp.id)[:8]); ws.cell(row=row, column=2, value=emp.name); ws.cell(row=row, column=3, value=emp.email)
        ws.cell(row=row, column=4, value=emp.role); ws.cell(row=row, column=5, value=p.dept.name if p and p.dept else "N/A")
        ws.cell(row=row, column=6, value=p.job_title if p else "N/A"); ws.cell(row=row, column=7, value=str(p.joining_date.date()) if p and p.joining_date else "N/A"); ws.cell(row=row, column=8, value=p.contact if p else "N/A")
    for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 18
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='payro-employees.xlsx')

@app.route('/export/payroll')
@role_required('admin', 'accounting')
def export_payroll():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Payroll Records"
    header_font = Font(bold=True, color="FFFFFF"); header_fill = PatternFill("solid", fgColor="6366F1")
    headers = ["Employee", "Email", "Month", "Year", "Net Amount", "Status", "Paid Date"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h); cell.font = header_font; cell.fill = header_fill
    for row, rec in enumerate(PayrollRecord.query.order_by(PayrollRecord.year.desc(), PayrollRecord.month.desc()).all(), 2):
        ws.cell(row=row, column=1, value=rec.user.name if rec.user else "N/A"); ws.cell(row=row, column=2, value=rec.user.email if rec.user else "N/A"); ws.cell(row=row, column=3, value=rec.month)
        ws.cell(row=row, column=4, value=rec.year); ws.cell(row=row, column=5, value=round(rec.net_amount, 2)); ws.cell(row=row, column=6, value=rec.status)
        ws.cell(row=row, column=7, value=str(rec.paid_date.date()) if rec.paid_date else "N/A")
    for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 16
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='payro-payroll.xlsx')

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 10000))
    print(f"DEBUG: App starting on port {port}...")
    print(f"DEBUG: DATABASE_URL present: {'Yes' if os.environ.get('DATABASE_URL') else 'No'}")
    print(f"DEBUG: SECRET_KEY present: {'Yes' if os.environ.get('SECRET_KEY') else 'No'}")
    app.run(host="0.0.0.0", port=port, debug=False)
